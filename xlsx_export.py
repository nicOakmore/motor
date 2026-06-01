"""
xlsx_export.py — Presupuesto in XLSX, plus a workbook with all the
secondary tables (plan de acopios, cuadro nº 2 descompuesto, banderas).

Spreadsheet-grade output for the office back-end (Excel / LibreOffice
/ Google Sheets). The PDF is for sending to the client; this is for
re-using the numbers downstream.
"""

from __future__ import annotations
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


BRAND = "1A3A5C"
BRAND_SOFT = "2D5D8D"
HEAD_BG = "EEF2F7"
GRAND_BG = "1A3A5C"
ROW_ALT = "FAFBFD"
MUTED = "5A6072"
THIN = Side(border_style="thin", color="D6DAE2")


def _border(cell):
    cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _header_cell(cell, label: str):
    cell.value = label
    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=BRAND, end_color=BRAND, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _border(cell)


def _money(v: float) -> float:
    return round(float(v or 0), 2)


def build_presupuesto_xlsx(out_path: pathlib.Path,
                            firm: dict,
                            meta: dict,
                            totales: dict,
                            partidas: list[dict],
                            acopios: list[dict],
                            capitulo_orden: list[str],
                            ref: str,
                            flags: list[dict] | None = None,
                            project_title: str | None = None) -> pathlib.Path:
    wb = Workbook()
    # --- Sheet 1: Presupuesto detallado por capítulo ---
    ws = wb.active
    ws.title = "Presupuesto"

    # Project header
    ws["A1"] = "PRESUPUESTO"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=BRAND)
    ws.merge_cells("A1:H1")
    ws["A2"] = f"Referencia: {ref}"
    ws["A2"].font = Font(name="Calibri", size=10, color=MUTED)
    ws.merge_cells("A2:H2")
    ws["A3"] = (f"{firm.get('nombre','')} · CIF {firm.get('cif','')} · "
                f"{firm.get('telefono','')} · {firm.get('email','')}")
    ws["A3"].font = Font(name="Calibri", size=9, color=MUTED)
    ws.merge_cells("A3:H3")

    # Project metadata block
    row = 5
    for k, v in (
        ("Promotor",      meta.get("promotor") or "—"),
        ("Emplazamiento", meta.get("emplazamiento") or "—"),
        ("Uso",           meta.get("uso") or "—"),
        ("Suelo",         meta.get("suelo") or "—"),
        ("Proyecto técnico", "Sí" if meta.get("requiere_proyecto") else "No"),
        ("Fecha",         project_title or ""),
    ):
        ws.cell(row=row, column=1, value=k).font = Font(name="Calibri", size=10, bold=True, color=BRAND)
        ws.cell(row=row, column=2, value=str(v)[:120]).font = Font(name="Calibri", size=10)
        ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=8)
        row += 1
    row += 1

    # Capítulo blocks
    by_cap: dict[str, list[dict]] = {}
    for p in partidas:
        by_cap.setdefault(p["capitulo"], []).append(p)
    ordered = [c for c in capitulo_orden if c in by_cap] + \
              [c for c in by_cap if c not in capitulo_orden]

    headers = ("Code", "Capítulo", "Descripción", "Ud", "Medición",
               "Precio ud", "Importe", "% PEM")
    for col, label in enumerate(headers, start=1):
        _header_cell(ws.cell(row=row, column=col), label)
    row += 1

    pem = float(totales.get("PEM") or 0)
    for cap in ordered:
        cap_partidas = by_cap[cap]
        cap_total = sum(float(p["importe"]) for p in cap_partidas)
        # Capítulo header row
        cell = ws.cell(row=row, column=1, value=cap[:5].upper())
        cell.font = Font(name="Calibri", size=10, bold=True, color=BRAND)
        cell.fill = PatternFill(start_color=HEAD_BG, end_color=HEAD_BG, fill_type="solid")
        cell2 = ws.cell(row=row, column=2, value=cap)
        cell2.font = Font(name="Calibri", size=10, bold=True, color=BRAND)
        cell2.fill = PatternFill(start_color=HEAD_BG, end_color=HEAD_BG, fill_type="solid")
        for c in range(3, 7):
            ws.cell(row=row, column=c).fill = PatternFill(
                start_color=HEAD_BG, end_color=HEAD_BG, fill_type="solid")
        ws.cell(row=row, column=7, value=_money(cap_total)).font = Font(
            name="Calibri", size=10, bold=True, color=BRAND)
        ws.cell(row=row, column=7).fill = PatternFill(
            start_color=HEAD_BG, end_color=HEAD_BG, fill_type="solid")
        ws.cell(row=row, column=7).number_format = '#,##0.00 €'
        row += 1
        # Partida rows
        for p in cap_partidas:
            ws.cell(row=row, column=1, value=p.get("code", ""))
            ws.cell(row=row, column=2, value=p.get("capitulo", ""))
            ws.cell(row=row, column=3, value=p.get("descripcion", ""))
            ws.cell(row=row, column=4, value=p.get("unidad", ""))
            ws.cell(row=row, column=5, value=float(p.get("medicion") or 0))
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=_money(p.get("precio_unitario")))
            ws.cell(row=row, column=6).number_format = '#,##0.00 €'
            imp = _money(p.get("importe"))
            ws.cell(row=row, column=7, value=imp)
            ws.cell(row=row, column=7).number_format = '#,##0.00 €'
            ws.cell(row=row, column=8, value=(imp / pem if pem else 0))
            ws.cell(row=row, column=8).number_format = '0.00%'
            row += 1
        row += 1

    # Cuadro resumen
    row += 1
    ws.cell(row=row, column=1, value="CUADRO RESUMEN (RD 1098/2001)").font = Font(
        name="Calibri", size=12, bold=True, color=BRAND)
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=8)
    row += 1
    resumen_rows = [
        ("PEM — Presupuesto de Ejecución Material", totales.get("PEM"), False),
        (f"Gastos Generales ({int(totales.get('gg_pct',0.13)*100)}%)",
         totales.get("GG"), False),
        (f"Beneficio Industrial ({int(totales.get('bi_pct',0.06)*100)}%)",
         totales.get("BI"), False),
        ("PEC — Presupuesto de Ejecución por Contrata",
         totales.get("PEC"), True),
    ]
    # IVA breakdown if mixed; otherwise single line
    iva_breakdown = totales.get("iva_breakdown") or []
    if iva_breakdown and len(iva_breakdown) > 1:
        for label, val in iva_breakdown:
            resumen_rows.append((label, val, False))
    else:
        resumen_rows.append(
            (f"IVA ({int(totales.get('iva_pct',0.10)*100)}%)",
             totales.get("IVA"), False))
    if totales.get("RETENCION"):
        resumen_rows.append((
            f"Retención IRPF ({(totales.get('retencion_pct',0)*100):.1f}%)",
            -float(totales.get("RETENCION")), False))
    if totales.get("RECARGO_EQUIV"):
        resumen_rows.append((
            f"Recargo de equivalencia ({(totales.get('recargo_pct',0)*100):.2f}%)",
            totales.get("RECARGO_EQUIV"), False))
    resumen_rows.append(("TOTAL", totales.get("TOTAL"), "grand"))
    resumen_rows.append(("ICIO (orientativo, sobre PEM)",
                         totales.get("ICIO"), "muted"))

    for label, val, kind in resumen_rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=7, value=_money(val))
        c2.number_format = '#,##0.00 €'
        if kind == "grand":
            c1.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            c2.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=GRAND_BG, end_color=GRAND_BG, fill_type="solid")
        elif kind is True:
            c1.font = Font(name="Calibri", size=10, bold=True, color=BRAND)
            c2.font = Font(name="Calibri", size=10, bold=True, color=BRAND)
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=HEAD_BG, end_color=HEAD_BG, fill_type="solid")
        elif kind == "muted":
            c1.font = Font(name="Calibri", size=10, italic=True, color=MUTED)
            c2.font = Font(name="Calibri", size=10, italic=True, color=MUTED)
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=6)
        row += 1

    # Column widths
    for col, width in zip("ABCDEFGH",
                          (10, 26, 56, 6, 10, 12, 14, 8)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A12"

    # --- Sheet 2: Plan de acopios ---
    if acopios:
        ws2 = wb.create_sheet("Plan de acopios")
        for col, label in enumerate(
                ("Material", "Descripción", "Ud", "Cantidad"), start=1):
            _header_cell(ws2.cell(row=1, column=col), label)
        for i, a in enumerate(acopios, start=2):
            ws2.cell(row=i, column=1, value=a.get("material_code", ""))
            ws2.cell(row=i, column=2, value=a.get("descripcion", ""))
            ws2.cell(row=i, column=3, value=a.get("unidad", ""))
            qty = ws2.cell(row=i, column=4, value=float(a.get("cantidad") or 0))
            qty.number_format = '#,##0.000'
        for col, width in zip("ABCD", (16, 50, 8, 12)):
            ws2.column_dimensions[col].width = width

    # --- Sheet 3: Descompuesto (cuadro Nº 2) ---
    rows_d = []
    for p in partidas:
        for c in p.get("descompuesto", []) or []:
            rows_d.append({
                "partida": p["code"],
                "partida_desc": p["descripcion"],
                "slot": c.get("slot", ""),
                "code": c.get("comp_code", ""),
                "descripcion": c.get("descripcion", ""),
                "unidad": c.get("unidad", ""),
                "rendimiento": c.get("rendimiento", 0),
                "precio_unitario": c.get("precio_unitario", 0),
                "importe": c.get("importe", 0),
            })
    if rows_d:
        ws3 = wb.create_sheet("Descompuesto")
        for col, label in enumerate(
                ("Partida", "Descripción partida", "Tipo", "Código",
                 "Componente", "Ud", "Rendim.", "Precio ud", "Importe"),
                start=1):
            _header_cell(ws3.cell(row=1, column=col), label)
        for i, r in enumerate(rows_d, start=2):
            ws3.cell(row=i, column=1, value=r["partida"])
            ws3.cell(row=i, column=2, value=r["partida_desc"])
            ws3.cell(row=i, column=3, value={
                "mo": "M.O.", "mat": "Mat.", "maq": "Maq."}.get(r["slot"], r["slot"]))
            ws3.cell(row=i, column=4, value=r["code"])
            ws3.cell(row=i, column=5, value=r["descripcion"])
            ws3.cell(row=i, column=6, value=r["unidad"])
            ws3.cell(row=i, column=7, value=float(r["rendimiento"])).number_format = '0.0000'
            ws3.cell(row=i, column=8, value=_money(r["precio_unitario"])).number_format = '#,##0.0000 €'
            ws3.cell(row=i, column=9, value=_money(r["importe"])).number_format = '#,##0.00 €'
        for col, width in zip("ABCDEFGHI",
                              (10, 44, 7, 10, 36, 6, 10, 14, 14)):
            ws3.column_dimensions[col].width = width

    # --- Sheet 4: Banderas regulatorias ---
    if flags:
        ws4 = wb.create_sheet("Banderas")
        for col, label in enumerate(
                ("Nivel", "Código", "Mensaje", "Regla origen"), start=1):
            _header_cell(ws4.cell(row=1, column=col), label)
        order = {"STOP": 0, "WARN": 1, "INFO": 2}
        for i, f in enumerate(
                sorted(flags, key=lambda x: order.get(x.get("nivel"), 9)),
                start=2):
            ws4.cell(row=i, column=1, value=f.get("nivel", ""))
            ws4.cell(row=i, column=2, value=f.get("codigo", ""))
            ws4.cell(row=i, column=3, value=f.get("mensaje", ""))
            ws4.cell(row=i, column=4, value=f.get("rule", ""))
        for col, width in zip("ABCD", (8, 24, 90, 26)):
            ws4.column_dimensions[col].width = width

    wb.save(str(out_path))
    return out_path
